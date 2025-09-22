import requests
import threading
import json
from openpyxl import load_workbook

def fetch_weather(kecamatan, result_dict):
    params = {
        "key": "4c0dbb9c648a4e669a641524251809",
        "q": kecamatan + ", Daerah Istimewa Yogyakarta, Indonesia",
        "lang" : "jv"
    }
    try:
        # Get Data Cuaca
        response = requests.get("http://api.weatherapi.com/v1/current.json", params=params, timeout=10)
        
        # Parse JSON response to object
        data = json.loads(response.text)

        # Simpan hasil ke dictionary
        result_dict[kecamatan] = [
            data["current"]["last_updated"],
            data["current"]["temp_c"],
            data["current"]["humidity"],
            data["current"]["condition"]["text"],
            data["current"]["wind_kph"],
            data["current"]["wind_dir"],
            data["current"]["uv"]
        ]

        # Debug output
        print([kecamatan] + result_dict[kecamatan])
    except Exception as e:
        print(f"{kecamatan}: ERROR {e}")
        result_dict[kecamatan] = ["ERROR"]*7

def main():
    # Load kecamatan from Excel
    wb = load_workbook('DIY.xlsx')
    ws = wb.active

    # Ambil daftar kecamatan dan simpan ke list
    kecamatan_list = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0]:
            kecamatan_list.append(row[0])
    

    # Dictionary untuk menyimpan hasil
    result_dict = {}

    # Simpan thread ke dalam list
    threads = []

    # Buat dan mulai thread untuk setiap kecamatan
    for kecamatan in kecamatan_list:
        t = threading.Thread(target=fetch_weather, args=(kecamatan, result_dict))
        threads.append(t)
        t.start()

    # Gabungkan semua thread
    for t in threads:
        t.join()

    # Header untuk hasil
    headers = ["Last Update", "Suhu (C)", "Kelembapan", "Kondisi Cuaca", "Kecepatan Angin (kph)", "Arah Angin", "Sinar UV"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=1, column=i, value=h)

    # Tulis hasil ke baris yang sesuai, mulai kolom B (kolom ke-2)
    for idx, kecamatan in enumerate(kecamatan_list, start=2):
        data = result_dict.get(kecamatan, ["-"]*7)
        for j, val in enumerate(data, start=2):
            ws.cell(row=idx, column=j, value=val)

    # Simpan ke excel
    
    wb.save('DIY.xlsx')

if __name__ == "__main__":
    main()